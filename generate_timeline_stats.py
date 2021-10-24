#! /usr/local/bin/python3
""" Generate reports with times between various events for various cohorts.

    A cohort consists of all students who apply to transfer to a college for a given term.

    reports/
      - Markdown report for each measure for each cohort
      - Spreadsheet for each measure for each report with multiple evaluations/registrations
        coalesced into single rows

    timelines/
      - Spreadsheet for each cohort showing all measures available per student

    ./
      - Baseline_Intervals_yyyy-mm-dd.xlsx Consolidated spreadsheet of statistics for each measure
        for each cohort.
      - cohort_report.txt Sizes of cohorts

    Event Dates:
      Session: Early Registration, Open Registration, Classes Start
      Admissions: Apply, Admit, Matric
      Registrations: First, Latest
      Transfers: First, Latest

    Algorithm:
      Get session event dates: these are fixed for all members of the cohort
      Get all students in the cohort
      For each student, the articulation term:
        First transfer fetch
        Latest transfer fetch
        First registration
        Latest registration

    Report, as of report date:
      Number of admitted; number matriculated
      Descriptive statistics: mean, std dev, median, mode, range, siqr for the following intervals:
        Apply to Admit
        Admit to Matric
        Matric to First Register
        Matric to Latest Register
        Admit to First Fetch
        Admit to Latest Fetch
        Matric to First Fetch
        Matric to Latest Fetch

    Not looked at, but potentially useful info:
      Students who belong to more than one transfer cohort for a semester. Does transfer fetch
      timing make them go to another college?

    Interesting to see how the above has evolved as the code below was developed.
"""

import sys
import argparse
import datetime
import statistics
import time

from collections import namedtuple, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from timeline_utils import min_sec
from pgconnection import PgConnection


class AdmitTerm:
  """ CF term code and semester name
  """
  def __init__(self, term_code, semester_name):
    self.term = int(term_code)
    self.name = semester_name

  def __repr__(self):
    return self.name


# Statistics
# -------------------------------------------------------------------------------------------------
""" Descriptive statistics for a cohort's measures
      stat_values[institution][admit_term][measure].n = 12345, etc
"""


class Stats:
  """ mean, median, mode, etc.
  """
  def __init__(self):
    self.n = self.mean = self.std_dev = self.median = self.mode = self.min_val = self.max_val =\
        self.q_1 = self.q_2 = self.q_3 = self.siqr = None


# Factory methods for initializing defaultdicts
def institution_factory():
  return defaultdict(term_factory)


def term_factory():
  return defaultdict(stat_factory)


def stat_factory():
  return Stats()


stat_values = defaultdict(institution_factory)

# App Parameters
# -------------------------------------------------------------------------------------------------
institution_names = {'BAR': 'Baruch', 'BCC': 'Bronx', 'BKL': 'Brooklyn', 'BMC': 'BMCC',
                     'CSI': 'Staten Island', 'CTY': 'City', 'HOS': 'Hostos', 'HTR': 'Hunter',
                     'JJC': 'John Jay', 'KCC': 'Kingsborough', 'LAG': 'LaGuardia', 'LEH': 'Lehman',
                     'MEC': 'Medgar Evers', 'NCC': 'Guttman', 'NYT': 'City Tech',
                     'QCC': 'Queensborough', 'QNS': 'Queens', 'SLU': 'Labor/Urban',
                     'SOJ': 'Journalism', 'SPH': 'Public Health', 'SPS': 'SPS', 'YRK': 'York'}

# Make the 4-letter admissions “program_action” names a little less terse/cryptic
action_to_event = {'appl': 'apply',
                   'admt': 'admit',
                   'dein': 'commit',
                   'matr': 'matric',
                   'wadm': 'admin'}

event_names = {'apply': 'Apply',
               'admit': 'Admit',
               'commit': 'Commit',
               'matric': 'Matric',
               'first_eval': 'First Evaluation',
               'latest_eval': 'Latest Evaluation',
               'start_early_enr': 'Start Early Enrollment',
               'start_open_enr': 'Start Open Enrollment',
               'first_reg': 'First Registration',
               'latest_reg': 'Latest Registration',
               'start_classes': 'Start Classes',
               'census_date': 'Census Date',
               'admin': 'Admin',
               }

# Admin "events" are included in the timelines spreadsheets for data verification; they can't be
# used for measurements
event_types = [key for key in event_names.keys() if key != 'admin']

EventPair = namedtuple('EventPair', 'earlier later')

# Where the evaluation posted_date was missing, I substituted January 1, 1901
missing_date = datetime.date(1901, 1, 1)


# events_dict()
# -------------------------------------------------------------------------------------------------
def events_dict():
  """ Factory method to produce default dates (None) for a student's events record.
  """
  events = {key: None for key in event_types}
  # Session info is same for all students in cohort
  events['start_early_enr'] = session.first_registration
  events['start_open_enr'] = session.start_open_registration
  events['start_classes'] = session.session_start
  events['census_date'] = session.census
  events['admin'] = []   # List of dein/wadm events with their dates
  return events


# Validate Command Line
# -------------------------------------------------------------------------------------------------
parser = argparse.ArgumentParser('Timelines by Cohort')
parser.add_argument('-t', '--admit_terms', type=int, nargs='*', default=[])
parser.add_argument('-i', '--institutions', nargs='*', default=[])
parser.add_argument('-e', '--event_pairs', nargs='*', default=[])
parser.add_argument('-d', '--debug', action='store_true')
args = parser.parse_args()

event_pairs = []
event_type_list = '\n  '.join([t for t in event_types if t != 'wadm'])
if len(args.event_pairs) < 1:
  print('NOTICE: no event pairs. No statistical reports will be produced.', file=sys.stderr)
for arg in args.event_pairs:
  try:
    earlier, later = arg.lower().split(':')
    if earlier in event_types and later in event_types:
      event_pairs.append(EventPair(earlier, later))
    else:
      raise ValueError('Unrecognized event_pair')
  except ValueError as ve:
    sys.exit(f'“{arg}” does not match earlier:later event_pair structure.\n'
             f'Valid event types are:\n  {event_type_list}')

if len(args.admit_terms) < 1 or len(args.institutions) < 1:
  sys.exit(f'Usage: -t admit_term... -i institution... -e event_pair...')

try:
  admit_terms = []
  for admit_term in args.admit_terms:
    year = 1900 + 100 * int(admit_term / 1000) + int(admit_term / 10) % 100
    assert year > 1989 and year < 2026, f'Admit Term year ({year}) must be between 1990 and 2025'
    month = admit_term % 10
    if month == 2:
      semester = 'Spring'
    elif month == 9:
      semester = 'Fall'
    else:
      sys.exit(f'{admit_term}: month ({month}) must be 2 for Spring or 9 for Fall.')
    semester = f'{semester} {year}'
    admit_terms.append(AdmitTerm(admit_term, semester))

except ValueError as ve:
  sys.exit(f'“{args.admit_term}” is not a valid CUNY term')

institutions = [i.strip('01').upper() for i in args.institutions]
for institution in institutions:
  if institution not in institution_names.keys():
    sys.exit(f'“{institution}” is not a valid CUNY institution')

conn = PgConnection('cuny_transfers')
cursor = conn.cursor()
cohort_report = open('./cohort_report.txt', 'w')

# Initialize Data Structures
# =================================================================================================
""" A cohort is a set of (students, institution, admit_term). Collect all 12 event dates for each
    cohort, then report each measure for each cohort.
"""
""" Generate separate reports in Markdown for each institution.
    Generate separate spreadsheets for each measure, with colleges as columns and statistical
    values as the rows. Preserve the order of the colleges from the command line.
"""
# Cohorts key is (institution, term), value is a dict with student_id as key, and dict of event
# dates as their values.
#   cohorts[(QNS, 1212)] = {12345678: {appl: 2020-10-10, admt: 2020-10-20, wadm: ...},
#                           87654321: {appl: 2020-11-11, admt: 2020-11-22, wadm: ...},
#                           ...}
start_time = time.time()
num_cohorts = len(admit_terms) * len(institutions)
print(f'Begin Generate Timeline Statistics\n  {len(event_pairs)} Event Pairs\n'
      f'  {num_cohorts} Cohorts', file=sys.stderr)
cohorts = dict()
for institution in institutions:
  for admit_term in sorted(admit_terms, key=lambda x: x.term):
    cohort_key = (institution, admit_term.term)
    student_ids = set()
    cohorts[cohort_key] = defaultdict(events_dict)

    # Get session events, which are the same for all students in the cohort
    # ---------------------------------------------------------------------------------------------
    Session = namedtuple('Session', 'institution term session first_registration '
                         'last_waitlist start_open_enristration last_registration '
                         'classes_start census_date sixty_percent classes_end')
    cursor.execute(f"""
        select * from sessions where institution = '{institution}' and term = {admit_term.term}
        """)
    session = None
    for row in cursor.fetchall():
      if session is None or row.session == '1':
        session = Session._make(row)
    if session is None:
      print(f'No session found for {institution_names[institution]} {admit_term}')
      continue

    # Add the students and their admission events to the cohort
    # ---------------------------------------------------------------------------------------------
    """ Merge Summer and Fall admit terms together here, and for evaluations and registrations.
        Students can only apply for Spring and Fall, but during the matriculation process, the Fall
        admit term gets changed to Summer so they can register then ... if they want to.
    """
    if (admit_term.term % 10) != 2:
      base_term = int(10 * (int(admit_term.term / 10)))
      term_clause = f'in ({base_term + 6}, {base_term + 9})'
    else:
      term_clause = f'= {admit_term.term}'

    cursor.execute(f"""
        select student_id, program_action, action_reason, effective_date from admissions
         where institution = '{institution}'
           and admit_term {term_clause}
           and program_action in ('APPL', 'ADMT', 'DEIN', 'MATR', 'WADM')
           order by effective_date
        """)
    for row in cursor.fetchall():
      student_ids.add(int(row.student_id))
      # For verificaton, show both commit (DEIN) and academic withdrawal (WADM) events as "Admin"
      program_action = row.program_action.lower()
      effective_date = row.effective_date
      if program_action in ['wadm', 'dein']:
        event_str = f'{row.program_action}:{row.action_reason}'.strip(':')
        cohorts[cohort_key][int(row.student_id)]['admin'].append(f'{effective_date} '
                                                                 f'{event_str}')
        # Any DEIN implies Commit
        cohorts[cohort_key][int(row.student_id)]['commit'] = row.effective_date
        # DEIN:ENDC and DEIN:DEPO imply Matric
        if event_str in ['DEIN:ENDC', 'DEIN:DEPO']:
          cohorts[cohort_key][int(row.student_id)]['matric'] = row.effective_date
      else:
        event_type = action_to_event[program_action]
        cohorts[cohort_key][int(row.student_id)][event_type] = effective_date

    print(f'{len(cohorts[cohort_key]):7,} students in {cohort_key} cohort', file=cohort_report)
    assert len(student_ids) == len(cohorts[cohort_key])
    student_id_list = ','.join(f'{id}' for id in student_ids)   # for looking up registrations

    # Transfer Evaluation dates
    # ---------------------------------------------------------------------------------------------
    if student_id_list != '':
      cursor.execute(f"""
        select student_id, posted_date
          from transfers_applied
         where dst_institution ~* '{institution}'
           and articulation_term {term_clause}
           and student_id in ({student_id_list})
      group by student_id, posted_date
        """)
      for row in cursor.fetchall():
        student_id = int(row.student_id)
        if (posted_date := row.posted_date) > datetime.date(1901, 1, 1):
          if cohorts[cohort_key][student_id]['first_eval'] is None \
             or posted_date < cohorts[cohort_key][student_id]['first_eval']:
            cohorts[cohort_key][student_id]['first_eval'] = posted_date
          if cohorts[cohort_key][student_id]['latest_eval'] is None \
             or posted_date > cohorts[cohort_key][student_id]['latest_eval']:
            cohorts[cohort_key][student_id]['latest_eval'] = posted_date

    # Enrollment dates
    # ---------------------------------------------------------------------------------------------
    if student_id_list != '':
      cursor.execute(f"""
        select student_id, first_date, last_date
          from registrations
         where institution ~* '{institution}'
           and term {term_clause}
           and student_id in ({student_id_list})
        """)
      for row in cursor.fetchall():
        if cohorts[cohort_key][row.student_id]['first_reg'] is None \
           or row.first_date < cohorts[cohort_key][row.student_id]['first_reg']:
          cohorts[cohort_key][row.student_id]['first_reg'] = row.first_date
        if cohorts[cohort_key][row.student_id]['latest_reg'] is None \
           or row.last_date > cohorts[cohort_key][row.student_id]['latest_reg']:
          cohorts[cohort_key][row.student_id]['latest_reg'] = row.last_date

    # Create a spreadsheet with the cohort's events for debugging/tableauing
    # ---------------------------------------------------------------------------------------------
    with open(f'./timelines/{institution}-{admit_term.term}.csv', 'w') as spreadsheet:
      print('Student ID,', ','.join([f'{event_names[name]}' for name in event_names.keys()]),
            file=spreadsheet)
      for student_id in sorted(student_ids):
        dates = ','.join([f'{cohorts[cohort_key][student_id][event]}' for event in event_types
                         if event != 'admin'])
        if len(cohorts[cohort_key][student_id]['admin']) == 0:
          dates += ',None'
        else:
          dates += ',' + '; '.join(sorted(cohorts[cohort_key][student_id]['admin']))
        print(f'{student_id}, {dates}', file=spreadsheet)

    # For each measure, generate a Markdown report, and collect stats for the measures spreadsheets
    # ---------------------------------------------------------------------------------------------
    for event_pair in event_pairs:
      s = stat_values[institution][admit_term.term][event_pair]
      earlier, later = event_pair
      with open(f'./reports/{institution}-{admit_term}-{earlier} to {later}.md', 'w') as report:
        print(f'# {institution_names[institution]}: {admit_term}\n\n'
              f'## Days from {event_names[earlier]} to {event_names[later]}\n'
              f'| Statistic | Value |\n| :--- | :--- |', file=report)
        # Build frequency distributions of earlier and later event date pair differences
        frequencies = defaultdict(int)  # Maybe plot these later
        deltas = []
        for student_id in cohorts[cohort_key].keys():
          if (cohorts[cohort_key][student_id][earlier] is not None
             and cohorts[cohort_key][student_id][earlier] != missing_date
             and cohorts[cohort_key][student_id][later] is not None
             and cohorts[cohort_key][student_id][later] != missing_date):
            delta = (cohorts[cohort_key][student_id][later]
                     - cohorts[cohort_key][student_id][earlier])

            deltas.append(delta.days)
            frequencies[delta.days] += 1

        s.n = len(deltas)
        print(f'| N | {s.n}', file=report)
        if len(deltas) > 5:
          s.mean = statistics.fmean(deltas)
          s.std_dev = statistics.stdev(deltas)
          s.median = statistics.median_grouped(deltas)
          s.mode = statistics.mode(deltas)
          s.min_val = min(deltas)
          s.max_val = max(deltas)
          min_max_str = f'{min(deltas)} : {max(deltas)}'
          quartile_list = statistics.quantiles(deltas, n=4, method='exclusive')
          s.q_1 = quartile_list[0]
          s.q_2 = quartile_list[1]
          s.q_3 = quartile_list[2]
          quartile_str = f'{s.q_1:.0f} : {s.q_2:.0f} : {s.q_3:.0f}'
          s.siqr = (s.q_3 - s.q_1) / 2.0
          print(f'| Medan | {s.median:.0f}', file=report)
          print(f'| Mean | {s.mean:.0f}', file=report)
          print(f'| Mode | {s.mode:.0f}', file=report)
          print(f'| Range | {min_max_str}', file=report)
          print(f'| Quartiles | {quartile_str}', file=report)
          print(f'| SIQR | {s.siqr:.1f}', file=report)
          print(f'| Std Dev | {s.std_dev:.1f}', file=report)
        else:
          print('### Not enough data.', file=report)

end_timelines = time.time()
print(f'That took {min_sec(end_timelines - start_time)}\nGenerate Workbook', file=sys.stderr)

# Generate an Excel workbook
# ------------------------------------------------------------------------------------------------
""" One sheet for each measure; colleges by columns; rows are statistics for admit term
"""
centered = Alignment('center')
bold = Font(bold=True)
wb = Workbook()
for event_pair in event_pairs:
  earlier, later = event_pair
  ws = wb.create_sheet(f'{event_names[earlier][0:14]} to {event_names[later][0:14]}')

  headings = [''] + institutions
  row = 1
  for col in range(len(headings)):
    ws.cell(row, col + 1, headings[col]).font = bold
    ws.cell(row, col + 1, headings[col]).alignment = centered

  for admit_term in admit_terms:
    row += 1
    ws.cell(row, 2, str(admit_term))
    ws.merge_cells(start_row=row, end_row=row, start_column=2, end_column=len(headings))
    ws.cell(row, 2).font = bold
    ws.cell(row, 2).alignment = centered

    # Everbody should have an N value
    row += 1
    ws.cell(row, 1, 'N').font = bold
    values = [stat_values[institution][admit_term.term][event_pair].n
              for institution in institutions]
    for col in range(2, 2 + len(headings) - 1):
      ws.cell(row, col).value = values[col - 2]

    # The remainder is messy because there will be None values where N < 6 for some institution, and
    # because different statistics have different formatting rules

    # Median
    row += 1
    ws.cell(row, 1, 'Median').font = bold
    values = []
    for institution in institutions:
      value = stat_values[institution][admit_term.term][event_pair].median
      if value is None:
        values.append('')
      else:
        values.append(value)
    for col in range(2, 2 + len(headings) - 1):
      ws.cell(row, col).value = values[col - 2]
      ws.cell(row, col).number_format = '0'
      ws.cell(row, col).font = bold

    # Mean
    row += 1
    ws.cell(row, 1, 'Mean').font = bold
    values = []
    for institution in institutions:
      value = stat_values[institution][admit_term.term][event_pair].mean
      if value is None:
        values.append('')
      else:
        values.append(value)
    for col in range(2, 2 + len(headings) - 1):
      ws.cell(row, col).value = values[col - 2]
      ws.cell(row, col).number_format = '0'

    # Mode
    row += 1
    ws.cell(row, 1, 'Mode').font = bold
    values = []
    for institution in institutions:
      value = stat_values[institution][admit_term.term][event_pair].mode
      if value is None:
        values.append('')
      else:
        values.append(value)
    for col in range(2, 2 + len(headings) - 1):
      ws.cell(row, col).value = values[col - 2]
      ws.cell(row, col).number_format = '0'

     # Min
    row += 1
    ws.cell(row, 1, 'Min').font = bold
    values = []
    for institution in institutions:
      value = stat_values[institution][admit_term.term][event_pair].min_val
      if value is None:
        values.append('')
      else:
        values.append(value)
    for col in range(2, 2 + len(headings) - 1):
      ws.cell(row, col).value = values[col - 2]
      ws.cell(row, col).number_format = '0'

     # Max
    row += 1
    ws.cell(row, 1, 'Max').font = bold
    values = []
    for institution in institutions:
      value = stat_values[institution][admit_term.term][event_pair].max_val
      if value is None:
        values.append('')
      else:
        values.append(value)
    for col in range(2, 2 + len(headings) - 1):
      ws.cell(row, col).value = values[col - 2]
      ws.cell(row, col).number_format = '0'

     # Q1
    row += 1
    ws.cell(row, 1, 'Q1').font = bold
    values = []
    for institution in institutions:
      value = stat_values[institution][admit_term.term][event_pair].q_1
      if value is None:
        values.append('')
      else:
        values.append(value)
    for col in range(2, 2 + len(headings) - 1):
      ws.cell(row, col).value = values[col - 2]
      ws.cell(row, col).number_format = '0.0'

     # Q2
    row += 1
    ws.cell(row, 1, 'Q2').font = bold
    values = []
    for institution in institutions:
      value = stat_values[institution][admit_term.term][event_pair].q_2
      if value is None:
        values.append('')
      else:
        values.append(value)
    for col in range(2, 2 + len(headings) - 1):
      ws.cell(row, col).value = values[col - 2]
      ws.cell(row, col).number_format = '0.0'

     # Q3
    row += 1
    ws.cell(row, 1, 'Q3').font = bold
    values = []
    for institution in institutions:
      value = stat_values[institution][admit_term.term][event_pair].q_3
      if value is None:
        values.append('')
      else:
        values.append(value)
    for col in range(2, 2 + len(headings) - 1):
      ws.cell(row, col).value = values[col - 2]
      ws.cell(row, col).number_format = '0.0'

     # SIQR
    row += 1
    ws.cell(row, 1, 'SIQR').font = bold
    values = []
    for institution in institutions:
      value = stat_values[institution][admit_term.term][event_pair].siqr
      if value is None:
        values.append('')
      else:
        values.append(value)
    for col in range(2, 2 + len(headings) - 1):
      ws.cell(row, col).value = values[col - 2]
      ws.cell(row, col).number_format = '0.1'

     # Std Dev
    row += 1
    ws.cell(row, 1, 'Std Dev').font = bold
    values = []
    for institution in institutions:
      value = stat_values[institution][admit_term.term][event_pair].std_dev
      if value is None:
        values.append('')
      else:
        values.append(value)
    for col in range(2, 2 + len(headings) - 1):
      ws.cell(row, col).value = values[col - 2]
      ws.cell(row, col).number_format = '0.1'

    # Empty row between Admit Terms
    row += 1
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=len(headings))

del wb['Sheet']
wb.save('./debug.xlsx')
end_time = time.time()
print(f'That took {min_sec(end_time - end_timelines)}\n'
      f'Generate Timeline Statistics took {min_sec(end_time - start_time)}')
