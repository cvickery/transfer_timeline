#! /usr/local/bin/python3
"""Generate reports with times between various events for various cohorts.

A cohort consists of all students who apply to transfer to a college for a given term.
There are two ways to build the cohorts:
  1.  Use the admissions table to find all students who applied for transfer admission to each
      college.
  2.  Use a “cohort” spreadsheet to get a list of emplids of interest, and use that to filter the
      first process.

  reports/
    - Markdown report for each measure for each cohort
    - Spreadsheet for each measure for each report with multiple evaluations/registrations
      coalesced into single rows

  timelines/
    - Spreadsheet for each cohort showing all measures available per student

  ./
    - Baseline_Intervals_yyyy-mm-dd.xlsx Consolidated spreadsheet of statistics for each measure
      for each cohort.
    - cohort_report.txt Sizes of cohorts

  Event Dates:
    Session: Early Registration, Open Registration, Classes Start
    Admissions: Apply, Admit, Matric
    Registrations: First, Latest
    Transfers: First, Latest

  Algorithm:
    Get session event dates: these are fixed for all members of the cohort
    Get all students in the cohort
    For each student, the articulation term:
      First transfer fetch
      Latest transfer fetch
      First registration
      Latest registration

  Report, as of report date:
    Number of admitted; number matriculated
    Descriptive statistics: mean, std dev, median, mode, range, siqr for the following intervals:
      Apply to Admit
      Admit to Matric
      Matric to First Register
      Matric to Latest Register
      Admit to First Fetch
      Admit to Latest Fetch
      Matric to First Fetch
      Matric to Latest Fetch

  Not looked at, but potentially useful info:
    Students who belong to more than one transfer cohort for a semester. Does transfer fetch
    timing make them go to another college?

  Interesting to see how the above has evolved as the code below was developed.
"""

import csv
import sys
import argparse
import datetime
import psycopg
import statistics
import time

from collections import namedtuple, defaultdict
from datetime import date
from math import sqrt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from pathlib import Path
from psycopg.rows import namedtuple_row
from subprocess import run
from timeline_utils import min_sec


class AdmitTerm:
  """CF term code and semester name."""

  def __init__(self, term_code, semester_name):
    """Capture the term_code and semester_name."""
    self.term = int(term_code)
    self.name = semester_name

  def __repr__(self):
    """Use the semester name as the representation of the object."""
    return self.name


# Statistics
# -------------------------------------------------------------------------------------------------
""" Descriptive statistics for a cohort's measures
      stat_values[institution][admit_term][event_pair].n = 12345, etc
"""


class Stats:
  """Descriptive statistics values (mean, median, mode, etc)."""

  def __init__(self):
    """Initialize as values as None."""
    self.n = self.mean = self.std_dev = self.median = self.mode = self.min_val = self.max_val =\
        self.q_1 = self.q_2 = self.q_3 = self.siqr = self.conf_int = None


# Factory methods for initializing defaultdicts
def institution_factory():
  """Create a defaultdict of term_factory defaultdicts."""
  return defaultdict(term_factory)


def term_factory():
  """Create a defaultdict of stat_factory objects."""
  return defaultdict(stat_factory)


def stat_factory():
  """Create a Stats object for a term_factory."""
  return Stats()


stat_values = defaultdict(institution_factory)

# App Parameters
# -------------------------------------------------------------------------------------------------
institution_names = {'BAR': 'Baruch', 'BCC': 'Bronx', 'BKL': 'Brooklyn', 'BMC': 'BMCC',
                     'CSI': 'Staten Island', 'CTY': 'City', 'HOS': 'Hostos', 'HTR': 'Hunter',
                     'JJC': 'John Jay', 'KCC': 'Kingsborough', 'LAG': 'LaGuardia', 'LEH': 'Lehman',
                     'MEC': 'Medgar Evers', 'NCC': 'Guttman', 'NYT': 'City Tech',
                     'QCC': 'Queensborough', 'QNS': 'Queens', 'SLU': 'Labor/Urban',
                     'SOJ': 'Journalism', 'SPH': 'Public Health', 'SPS': 'SPS', 'YRK': 'York'}

# Make the 4-letter admissions “program_action” names a little less terse/cryptic
action_to_event = {'appl': 'apply',
                   'admt': 'admit',
                   'dein': 'commit',
                   'matr': 'matric',
                   'wadm': 'admin'}

event_names = {'apply': 'Apply',
               'admit': 'Admit',
               'commit': 'Commit',
               'matric': 'Matric',
               'first_eval': 'First Eval',
               'latest_eval': 'Latest Eval',
               'start_early_enr': 'Early Enroll',
               'start_open_enr': 'Open Enroll',
               'start_classes': 'Start Classes',
               'census_date': 'Census Date',
               'first_reg': 'First Register',
               'latest_reg': 'Latest Register',
               'admin': 'Admin',
               }

event_definitions = {'EVENT NAME': 'DEFINITION (command line code)',
                     'Apply': 'Student submitted transfer application (apply)',
                     'Admit': 'College admitted student (admit)',
                     'Commit': 'Student committed to attend (commit)',
                     'Matric': 'Student matriculated (matric)',
                     'First Eval': 'First date college evaluated student’s courses (first_eval)',
                     'Latest Eval': 'Latest date college evaluated student’s courses (latest_eval)',
                     'Early Enroll': 'Start of early enrollment period (start_early_enr)',
                     'Open Enroll': 'Start of open enrollment period (start_open_enr)',
                     'Start Classes': 'First day of classes (start_classes)',
                     'Census Date': 'Official enrollment headcount date (census_date)',
                     'First Register': 'Date student first registered for courses (first_reg)',
                     'Latest Register': 'Latest date student altered registration (latest_reg)',
                     }

# Admin "events" are included in the timelines spreadsheets for data verification; they can't be
# used for measurements
event_types = [key for key in event_names.keys() if key != 'admin']

EventPair = namedtuple('EventPair', 'earlier later')

# Where the evaluation posted_date was missing, I substituted January 1, 1901
missing_date = datetime.date(1901, 1, 1)


# events_dict()
# -------------------------------------------------------------------------------------------------
def events_dict():
  """Produce default dates (None) for a student’s events record."""
  events = {key: None for key in event_types}

  # Session info is the same for all students in a cohort, so that is initialized here.
  events['start_early_enr'] = session.early_enrollment
  events['start_open_enr'] = session.open_enrollment
  events['start_classes'] = session.classes_start
  events['census_date'] = session.census_date
  events['admin'] = []   # List of dein/wadm events with their dates
  return events


# Available terms and sessions cache
# -------------------------------------------------------------------------------------------------
with psycopg.connect('dbname=cuny_transfers') as conn:
  with conn.cursor(row_factory=namedtuple_row) as cursor:
    cursor.execute("""select count(*), term
                        from sessions
                       where term >=1132
                         and term::text ~* '[29]$'
                       group by term
                       order by term;
                    """)
    available_terms = [str(row.term) for row in cursor]
    cursor.execute("""select *
                        from sessions
                       where session='1'
                         and term >=1132
                         and term::text ~* '[29]$'
                       order by institution, term
                    """)
    sessions_cache = {(row.institution[0:3], row.term): row for row in cursor}


# Validate Command Line
# -------------------------------------------------------------------------------------------------
parser = argparse.ArgumentParser('Timelines by Cohort')
parser.add_argument('-t', '--admit_terms', nargs='*', default=available_terms)
parser.add_argument('-i', '--institutions', nargs='*', default=['bcc', 'bmc', 'hos', 'kcc', 'lag',
                                                                'qcc', 'csi', 'mec', 'nyt', 'bar',
                                                                'bkl', 'cty', 'htr', 'jjc', 'leh',
                                                                'qns', 'sps', 'yrk'])
parser.add_argument('-e', '--event_pairs', nargs='*', default=['apply:admit',
                                                               'admit:commit',
                                                               'commit:matric',
                                                               'admit:matric',
                                                               'admit:first_eval',
                                                               'admit:latest_eval',
                                                               'admit:start_open_enr',
                                                               'commit:first_eval',
                                                               'commit:latest_eval',
                                                               'matric:first_eval',
                                                               'matric:latest_eval',
                                                               'first_eval:start_open_enr',
                                                               'latest_eval:start_open_enr',
                                                               'first_eval:start_classes',
                                                               'latest_eval:start_classes',
                                                               'first_eval:census_date',
                                                               'latest_eval:census_date'])
parser.add_argument('-esc', '--explicit_student_cohort')
parser.add_argument('-d', '--debug', action='store_true')
parser.add_argument('-n', '--event_names', action='store_true')
parser.add_argument('-s', '--stats', nargs='*', default=['n',
                                                         'median',
                                                         'mean',
                                                         'mode',
                                                         'min',
                                                         'max',
                                                         'q1',
                                                         'q2',
                                                         'q3',
                                                         'siqr',
                                                         'std_dev'])
parser.add_argument('-nop', '--no_progress', action='store_true')
args = parser.parse_args()

show_progress = not args.no_progress

# If event names are requested, show the possibilities and exit.
if args.event_names:
  print('            ')
  for k, v in event_definitions.items():
    print(f'{ k:16} {v}')
  exit('')

# Handle explicit student cohort list, if present.
explicit_student_cohort = []
if args.explicit_student_cohort:
  with open(args.explicit_student_cohort, 'r') as esc_file:
    reader = csv.reader(esc_file)
    for line in reader:
      if reader.line_num == 1:
        Row = namedtuple('Row', [col.lower().replace(' ', '_') for col in line])
        if 'empl_id' not in Row._fields:
          exit('Explicit Student Cohort file has no “empl_id” column')
      else:
        row = Row._make(line)
        explicit_student_cohort.append(f'{int(row.empl_id)}')
  explicit_student_cohort_clause = f'and student_id in ({",".join(explicit_student_cohort)})'
else:
  explicit_student_cohort_clause = ''

# Be sure queries/ file set is consistent
is_copacetic = run(['./check_queries.py', '-nop'])
if is_copacetic.returncode:
  exit('Query check failed')

# Process processing options
stats_to_show = [stat for stat in args.stats]
if len(stats_to_show) < 1:
  exit('No stats to show')

if len(args.event_pairs) < 1:
  exit('No event pairs')

event_type_list = '\n  '.join([t for t in event_types if t != 'wadm'])
event_pairs = []

for arg in args.event_pairs:
  try:
    earlier, later = arg.lower().split(':')
    if earlier in event_types and later in event_types:
      event_pairs.append(EventPair(earlier, later))
    else:
      raise ValueError('Unrecognized event_pair')
  except ValueError:
    exit(f'“{arg}” does not match earlier:later event_pair structure.\n'
         f'Valid event types are:\n  {event_type_list}')

if len(args.admit_terms) < 1 or len(args.institutions) < 1:
  sys.exit('Usage: -t admit_term... -i institution... -e event_pair...')

admit_terms = []
for admit_term in args.admit_terms:
  if admit_term not in available_terms:
    available_terms_str = ', '.join(available_terms)
    exit(f'{admit_term} is not one of: {available_terms_str}')
  admit_term = int(admit_term)
  year = 1900 + 100 * int(admit_term / 1000) + int(admit_term / 10) % 100
  month = admit_term % 10
  semester = 'Spring' if month == 2 else 'Fall'
  semester = f'{semester} {year}'
  admit_terms.append(AdmitTerm(admit_term, semester))

# Senior colleges for "super cohort"
senior_colleges = ['BAR', 'BKL', 'CTY', 'HTR', 'JJC', 'LEH', 'QNS', 'SLU', 'SPS', 'YRK']
# Column heading name for super_cohort colleges, with repeated letters removed ('BCHJLQSY')
super_cohort = ''.join([sc[0] for sc in senior_colleges]).replace('BB', 'B').replace('SS', 'S')

# Institutions to show, in left to right order (from command line)
institutions = [i.strip('01').upper() for i in args.institutions]
for institution in institutions:
  if institution not in institution_names.keys():
    sys.exit(f'“{institution}” is not a valid CUNY institution')
institutions_str = ','.join([f"'{institution}01'" for institution in institutions])

conn = psycopg.connect('dbname=cuny_transfers')
cursor = conn.cursor(row_factory=namedtuple_row)
cohort_report = open('./cohort_report.txt', 'w')

# Initialize Data Structures
# =================================================================================================
""" A cohort is a set of (students, institution, admit_term). Collect all 12 event dates for each
    cohort, then report each measure for each cohort.
"""
""" Generate separate reports in Markdown for each institution.
    Generate separate spreadsheets for each measure, with colleges as columns and statistical
    values as the rows. Preserve the order of the colleges from the command line.
"""
# Cohorts key is (institution, term), value is a dict with student_id as key, and dict of event
# dates as their values.
#   cohorts[(QNS, 1212)] = {12345678: {appl: 2020-10-10, admt: 2020-10-20, wadm: ...},
#                           87654321: {appl: 2020-11-11, admt: 2020-11-22, wadm: ...},
#                           ...}
start_time = time.time()
num_cohorts = len(admit_terms) * (len(institutions) + 1)
print(f'Begin Generate Timeline Statistics\n  {len(event_pairs)} Event Pairs\n'
      f'  {len(admit_terms)} terms × {len(institutions)} institutions => {num_cohorts} Cohorts',
      file=sys.stderr)

cohorts = dict()
super_cohort_deltas = defaultdict(list)

cohort_num = 0
for institution in institutions:
  for admit_term in sorted(admit_terms, key=lambda x: x.term):
    cohort_num += 1
    if show_progress:
      print(f'\rCohort {cohort_num:,}/{num_cohorts:,}', end='')

    student_ids = set()

    cohort_key = (institution, admit_term.term)
    cohorts[cohort_key] = defaultdict(events_dict)

    super_cohort_key = (super_cohort, admit_term.term)
    cohorts[super_cohort_key] = defaultdict(events_dict)

    # Get session events for the cohort (and super_cohort)
    try:
      session = sessions_cache[(institution, admit_term.term)]
    except KeyError:
      # No session for this admit_term for this institution (yet)
      print(f'\nNo session for {institution} {admit_term.term}')
      continue

    # Add the students and their admission events to the cohort
    # ---------------------------------------------------------------------------------------------
    """ Merge Summer and Fall admit terms together here, and for evaluations and registrations.
        Students can only apply for Spring and Fall, but during the matriculation process, the Fall
        admit term gets changed to Summer so they can register then ... if they want to.
    """
    if (admit_term.term % 10) != 2:
      base_term = int(10 * (int(admit_term.term / 10)))
      term_clause = f'in ({base_term + 6}, {base_term + 9})'
    else:
      term_clause = f'= {admit_term.term}'

    cursor.execute(f"""
        select student_id, program_action, action_reason, effective_date
          from admissions
         where institution = '{institution}01'
           {explicit_student_cohort_clause}
           and admit_term {term_clause}
           and program_action in ('APPL', 'ADMT', 'DEIN', 'MATR', 'WADM')
           order by effective_date
        """)

    if args.debug:
      print(f'{institution} {term_clause} has {cursor.rowcount:,} admission events')

    for row in cursor.fetchall():
      student_ids.add(int(row.student_id))

      program_action = row.program_action.lower()
      effective_date = row.effective_date

      # For verificaton, show both commit (DEIN) and academic withdrawal (WADM) events as "Admin"
      if program_action in ['wadm', 'dein']:
        event_str = f'{row.program_action}:{row.action_reason}'.strip(':')
        cohorts[cohort_key][int(row.student_id)]['admin'].append(f'{effective_date} '
                                                                 f'{event_str}')
        cohorts[super_cohort_key][int(row.student_id)]['admin'].append(f'{effective_date} '
                                                                       f'{event_str}')
        # Any DEIN implies Commit
        if program_action == 'dein':
          cohorts[cohort_key][int(row.student_id)]['commit'] = effective_date
          cohorts[super_cohort_key][int(row.student_id)]['commit'] = effective_date
          # DEIN:ENDC and DEIN:DEPO imply Matric
          if event_str in ['DEIN:ENDC', 'DEIN:DEPO']:
            cohorts[cohort_key][int(row.student_id)]['matric'] = effective_date
            cohorts[super_cohort_key][int(row.student_id)]['matric'] = effective_date

      else:
        event_type = action_to_event[program_action]
        cohorts[cohort_key][int(row.student_id)][event_type] = effective_date
        cohorts[super_cohort_key][int(row.student_id)][event_type] = effective_date

    print(f'{len(cohorts[cohort_key]):7,} students in {cohort_key} cohort', file=cohort_report)
    assert len(student_ids) == len(cohorts[cohort_key])
    student_id_list = ','.join(f'{id}' for id in student_ids)   # for looking up registrations

    # Transfer Evaluation dates
    # ---------------------------------------------------------------------------------------------
    if student_id_list != '':
      cursor.execute(f"""
        select student_id, posted_date
          from transfers_applied
         where dst_institution ~* '{institution}'
           and articulation_term {term_clause}
           and student_id in ({student_id_list})
      group by student_id, posted_date
        """)
      for row in cursor.fetchall():
        student_id = int(row.student_id)
        if (posted_date := row.posted_date) > datetime.date(1901, 1, 1):
          if cohorts[cohort_key][student_id]['first_eval'] is None \
             or posted_date < cohorts[cohort_key][student_id]['first_eval']:
            cohorts[cohort_key][student_id]['first_eval'] = posted_date
            cohorts[super_cohort_key][student_id]['first_eval'] = posted_date
          if cohorts[cohort_key][student_id]['latest_eval'] is None \
             or posted_date > cohorts[cohort_key][student_id]['latest_eval']:
            cohorts[cohort_key][student_id]['latest_eval'] = posted_date
            cohorts[super_cohort_key][student_id]['latest_eval'] = posted_date

    # Registration dates
    # ---------------------------------------------------------------------------------------------
    # Although we collect drop dates, we report only first and last add dates (for now).
    if student_id_list != '':
      cursor.execute(f"""
      select student_id, min(add_date) as first_add,
                         max(add_date) as last_add,
                         count(add_date) as num_adds,
                         min(drop_date) as first_drop,
                         max(drop_date) as last_drop,
                         count(drop_date) as num_drops
      from registrations
      where institution = '{institution}01'
      and term {term_clause}
      and student_id in ({student_id_list})
      group by institution, student_id
        """)
      for row in cursor.fetchall():

        cohorts[cohort_key][row.student_id]['first_reg'] = row.first_add
        cohorts[super_cohort_key][row.student_id]['first_reg'] = row.first_add

        cohorts[cohort_key][row.student_id]['latest_reg'] = row.last_add
        cohorts[super_cohort_key][row.student_id]['latest_reg'] = row.last_add

    # Create a spreadsheet with the cohort's events for debugging/tableau-ing/powerbi-ing
    # ---------------------------------------------------------------------------------------------
    # Super cohort not included here
    with open(f'./timelines/{institution}-{admit_term.term}.csv', 'w') as spreadsheet:
      print('Student ID,', ','.join([f'{event_names[name]}' for name in event_names.keys()]),
            file=spreadsheet)
      for student_id in sorted(student_ids):
        dates = ','.join([f'{cohorts[cohort_key][student_id][event]}' for event in event_types
                         if event != 'admin'])
        if len(cohorts[cohort_key][student_id]['admin']) == 0:
          dates += ',None'
        else:
          dates += ',' + '; '.join(sorted(cohorts[cohort_key][student_id]['admin']))
        print(f'{student_id}, {dates}', file=spreadsheet)

    # For each measure, generate a Markdown report, and collect stats for the measures workbook
    # ---------------------------------------------------------------------------------------------
    for event_pair in event_pairs:
      s = stat_values[institution][admit_term.term][event_pair]

      earlier, later = event_pair
      super_cohort_event_key = super_cohort_key + (event_pair, )
      with open(f'./reports/{institution}-{admit_term}-{earlier} to {later}.md', 'w') as report:
        print(f'# {institution_names[institution]}: {admit_term}\n\n'
              f'## Days from {event_names[earlier]} to {event_names[later]}\n'
              f'| Statistic | Value |\n| :--- | :--- |', file=report)
        # Build frequency distributions of earlier and later event date pair differences
        frequencies = defaultdict(int)  # Maybe plot these later
        deltas = []
        for student_id in cohorts[cohort_key].keys():
          if (cohorts[cohort_key][student_id][earlier] is not None
             and cohorts[cohort_key][student_id][earlier] != missing_date
             and cohorts[cohort_key][student_id][later] is not None
             and cohorts[cohort_key][student_id][later] != missing_date):
            try:
              delta = (cohorts[cohort_key][student_id][later]
                       - cohorts[cohort_key][student_id][earlier])
            except TypeError:
              print(f'{cohort_key=} {student_id=} {earlier=} {later=} '
                    f'{cohorts[cohort_key][student_id][earlier]=} '
                    f'{cohorts[cohort_key][student_id][later]=}')
              exit()
            deltas.append(delta.days)
            frequencies[delta.days] += 1
            if institution in senior_colleges:
              super_cohort_deltas[super_cohort_event_key].append(delta.days)

        s.n = len(deltas)
        print(f'| N | {s.n}', file=report)
        if len(deltas) > 5:
          s.mean = statistics.fmean(deltas)
          s.std_dev = statistics.stdev(deltas)
          s.median = statistics.median_grouped(deltas)
          s.mode = statistics.mode(deltas)
          s.min_val = min(deltas)
          s.max_val = max(deltas)
          min_max_str = f'{min(deltas)} : {max(deltas)}'
          quartile_list = statistics.quantiles(deltas, n=4, method='exclusive')
          s.q_1 = quartile_list[0]
          s.q_2 = quartile_list[1]
          s.q_3 = quartile_list[2]
          quartile_str = f'{s.q_1:.0f} : {s.q_2:.0f} : {s.q_3:.0f}'
          s.siqr = (s.q_3 - s.q_1) / 2.0
          print(f'| Medan | {s.median:.0f}', file=report)
          print(f'| Mean | {s.mean:.0f}', file=report)
          print(f'| Mode | {s.mode:.0f}', file=report)
          print(f'| Range | {min_max_str}', file=report)
          print(f'| Quartiles | {quartile_str}', file=report)
          print(f'| SIQR | {s.siqr:.1f}', file=report)
          print(f'| Std Dev | {s.std_dev:.1f}', file=report)
        else:
          print('### Not enough data.', file=report)

# Calculate statistics for the super cohort.
print('\nCalculate Statistics', file=sys.stderr)
for admit_term in admit_terms:
  for event_pair in event_pairs:
    s = stat_values[super_cohort][admit_term.term][event_pair]

    earlier, later = event_pair
    deltas = super_cohort_deltas[(super_cohort, admit_term.term, event_pair)]

    s.n = len(deltas)
    if len(deltas) > 5:
      s.mean = statistics.fmean(deltas)
      s.std_dev = statistics.stdev(deltas)
      s.median = statistics.median_grouped(deltas)
      s.mode = statistics.mode(deltas)
      s.min_val = min(deltas)
      s.max_val = max(deltas)
      min_max_str = f'{min(deltas)} : {max(deltas)}'
      quartile_list = statistics.quantiles(deltas, n=4, method='exclusive')
      s.q_1 = quartile_list[0]
      s.q_2 = quartile_list[1]
      s.q_3 = quartile_list[2]
      s.siqr = (s.q_3 - s.q_1) / 2.0


# Write statistics to db
# ------------------------------------------------------------------------------------------------
print('Write statistics to db')

# All query files should have the same date (via check_queries.py), so get it for student_summary
files_date = date.fromtimestamp(Path('./queries/CV_QNS_STUDENT_SUMMARY.csv').stat().st_ctime)
with psycopg.connect('dbname=cuny_transfers') as conn:
  with conn.cursor() as cursor:
    cursor.execute('delete from statistics')
    cursor.execute('delete from statistics_dates')
    cursor.execute('insert into statistics_dates values(%s, %s)', (files_date, date.today()))
    for event_pair in event_pairs:
      for admit_term in admit_terms:
        if 0 == stat_values[super_cohort][admit_term.term][event_pair].n:
          # Skip terms where there is no data yet
          continue
        for institution in institutions + [super_cohort]:
          values = [institution, admit_term.term, event_pair]

          n = stat_values[institution][admit_term.term][event_pair].n
          std_dev = stat_values[institution][admit_term.term][event_pair].std_dev
          if n > 0 and std_dev is not None:
            conf_95 = 0.95 * (std_dev / sqrt(n))
          else:
            conf_95 = None

          values.append(n)
          values.append(stat_values[institution][admit_term.term][event_pair].median)
          values.append(stat_values[institution][admit_term.term][event_pair].siqr)
          values.append(stat_values[institution][admit_term.term][event_pair].mean)
          values.append(std_dev)
          values.append(conf_95)
          values.append(stat_values[institution][admit_term.term][event_pair].mode)
          values.append(stat_values[institution][admit_term.term][event_pair].min_val)
          values.append(stat_values[institution][admit_term.term][event_pair].max_val)
          values.append(stat_values[institution][admit_term.term][event_pair].q_1)
          values.append(stat_values[institution][admit_term.term][event_pair].q_2)
          values.append(stat_values[institution][admit_term.term][event_pair].q_3)

          cursor.execute("""
          insert into statistics
                 values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", values)


# Generate Excel workbook
# ------------------------------------------------------------------------------------------------
print('Generate Workbook', file=sys.stderr)
""" One sheet for each measure; colleges by columns; rows are statistics for admit term
"""
centered = Alignment('center')
bold = Font(bold=True)
wb = Workbook()
institutions.append(super_cohort)
for event_pair in event_pairs:
  earlier, later = event_pair
  ws = wb.create_sheet(f'{event_names[earlier][0:14]} to {event_names[later][0:14]}')

  headings = [''] + institutions
  row = 1
  for col in range(len(headings)):
    ws.cell(row, col + 1, headings[col]).font = bold
    ws.cell(row, col + 1, headings[col]).alignment = centered

  for admit_term in admit_terms:

    # Skip terms for which there are no data yet (i.e., super cohort N is zero)
    if 0 == stat_values[institutions[-1]][admit_term.term][event_pair].n:
      continue

    row += 1
    ws.cell(row, 2, str(admit_term))
    ws.merge_cells(start_row=row, end_row=row, start_column=2, end_column=len(headings))
    ws.cell(row, 2).font = bold
    ws.cell(row, 2).alignment = centered

    # Everybody should have an N value
    row += 1
    ws.cell(row, 1, 'N').font = bold
    values = [stat_values[institution][admit_term.term][event_pair].n
              for institution in institutions]
    for col in range(2, 2 + len(headings) - 1):
      ws.cell(row, col).value = values[col - 2]

    # The remainder is messy because there will be None values where N < 6 for some institution,
    # and because different statistics have different formatting rules

    # Median
    if 'median' in stats_to_show:
      row += 1
      ws.cell(row, 1, 'Median').font = bold
      values = []
      for institution in institutions:
        value = stat_values[institution][admit_term.term][event_pair].median
        if value is None:
          values.append('')
        else:
          values.append(value)
      for col in range(2, 2 + len(headings) - 1):
        ws.cell(row, col).value = values[col - 2]
        ws.cell(row, col).number_format = '0.0'
        ws.cell(row, col).font = bold

    # SIQR
    if 'siqr' in stats_to_show:
      row += 1
      ws.cell(row, 1, 'SIQR').font = bold
      values = []
      for institution in institutions:
        value = stat_values[institution][admit_term.term][event_pair].siqr
        if value is None:
          values.append('')
        else:
          values.append(value)
      for col in range(2, 2 + len(headings) - 1):
        ws.cell(row, col).value = values[col - 2]
        ws.cell(row, col).number_format = '0.0'

    # Mean
    if 'mean' in stats_to_show:
      row += 1
      ws.cell(row, 1, 'Mean').font = bold
      values = []
      for institution in institutions:
        value = stat_values[institution][admit_term.term][event_pair].mean
        if value is None:
          values.append('')
        else:
          values.append(value)
      for col in range(2, 2 + len(headings) - 1):
        ws.cell(row, col).value = values[col - 2]
        ws.cell(row, col).number_format = '0.0'

    # Std Dev
    if 'std_dev' in stats_to_show:
      row += 1
      ws.cell(row, 1, 'Std Dev').font = bold
      values = []
      for institution in institutions:
        value = stat_values[institution][admit_term.term][event_pair].std_dev
        if value is None:
          values.append('')
        else:
          values.append(value)
      for col in range(2, 2 + len(headings) - 1):
        ws.cell(row, col).value = values[col - 2]
        ws.cell(row, col).number_format = '0.0'

      # 95% Confidence interval (derived from std dev and n); always shown with std_dev
      row += 1
      ws.cell(row, 1, '95% Conf').font = bold
      values = []
      for institution in institutions:
        n = stat_values[institution][admit_term.term][event_pair].n
        std_dev = stat_values[institution][admit_term.term][event_pair].std_dev
        if n and n > 0 and std_dev is not None:
          value = 0.95 * (std_dev / sqrt(n))
          values.append(value)
        else:
          values.append('')
      for col in range(2, 2 + len(headings) - 1):
        ws.cell(row, col).value = values[col - 2]
        ws.cell(row, col).number_format = '0.00'

    # Mode
    if 'mode' in stats_to_show:
      row += 1
      ws.cell(row, 1, 'Mode').font = bold
      values = []
      for institution in institutions:
        value = stat_values[institution][admit_term.term][event_pair].mode
        if value is None:
          values.append('')
        else:
          values.append(value)
      for col in range(2, 2 + len(headings) - 1):
        ws.cell(row, col).value = values[col - 2]
        ws.cell(row, col).number_format = '0'

    # Min
    if 'min' in stats_to_show:
      row += 1
      ws.cell(row, 1, 'Min').font = bold
      values = []
      for institution in institutions:
        value = stat_values[institution][admit_term.term][event_pair].min_val
        if value is None:
          values.append('')
        else:
          values.append(value)
      for col in range(2, 2 + len(headings) - 1):
        ws.cell(row, col).value = values[col - 2]
        ws.cell(row, col).number_format = '0'

    # Max
    if 'max' in stats_to_show:
      row += 1
      ws.cell(row, 1, 'Max').font = bold
      values = []
      for institution in institutions:
        value = stat_values[institution][admit_term.term][event_pair].max_val
        if value is None:
          values.append('')
        else:
          values.append(value)
      for col in range(2, 2 + len(headings) - 1):
        ws.cell(row, col).value = values[col - 2]
        ws.cell(row, col).number_format = '0'

    # Q1
    if 'q1' in stats_to_show:
      row += 1
      ws.cell(row, 1, 'Q1').font = bold
      values = []
      for institution in institutions:
        value = stat_values[institution][admit_term.term][event_pair].q_1
        if value is None:
          values.append('')
        else:
          values.append(value)
      for col in range(2, 2 + len(headings) - 1):
        ws.cell(row, col).value = values[col - 2]
        ws.cell(row, col).number_format = '0.0'

    # Q2
    if 'q2' in stats_to_show:
      row += 1
      ws.cell(row, 1, 'Q2').font = bold
      values = []
      for institution in institutions:
        value = stat_values[institution][admit_term.term][event_pair].q_2
        if value is None:
          values.append('')
        else:
          values.append(value)
      for col in range(2, 2 + len(headings) - 1):
        ws.cell(row, col).value = values[col - 2]
        ws.cell(row, col).number_format = '0.0'

    # Q3
    if 'q3' in stats_to_show:
      row += 1
      ws.cell(row, 1, 'Q3').font = bold
      values = []
      for institution in institutions:
        value = stat_values[institution][admit_term.term][event_pair].q_3
        if value is None:
          values.append('')
        else:
          values.append(value)
      for col in range(2, 2 + len(headings) - 1):
        ws.cell(row, col).value = values[col - 2]
        ws.cell(row, col).number_format = '0.0'

    # Empty row between Admit Terms
    row += 1
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=len(headings))

del wb['Sheet']
wb.save(f'./xlsx_archive/{date.today()}.xlsx')

print(f'Total Time {min_sec(time.time() - start_time)}')
